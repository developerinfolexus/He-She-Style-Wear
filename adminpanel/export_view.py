from django.http import HttpResponse
from django.db.models import Q
from decimal import Decimal
import datetime
from django.utils import timezone
from customer.models import Order, CheckoutDetails
from django.views.decorators.csrf import csrf_exempt
from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required(login_url='admin_login')
def export_orders_excel(request):
    """
    Export orders to Excel with filtering logic reused from dashboard.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    # Reuse filtering logic from dashboard
    selected_country = request.GET.get('country')
    date_filter = request.GET.get('date_filter', '')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    # Build base queryset from CheckoutDetails
    qs = CheckoutDetails.objects.all()
    if selected_country:
        sc = str(selected_country).strip().upper()
        if sc == 'US':
            variants = ['US', 'USA', 'United States', 'United States of America']
            qs = qs.filter(country__in=variants)
        elif sc == 'CA':
            variants = ['CA', 'CAN', 'Canada']
            qs = qs.filter(country__in=variants)

    # Date filtering
    start_dt = None
    end_dt = None
    now = timezone.now()
    if date_filter == 'today':
        start_dt = timezone.make_aware(datetime.datetime.combine(now.date(), datetime.time.min))
        end_dt = timezone.make_aware(datetime.datetime.combine(now.date(), datetime.time.max))
    elif date_filter == 'yesterday':
        yesterday = now.date() - datetime.timedelta(days=1)
        start_dt = timezone.make_aware(datetime.datetime.combine(yesterday, datetime.time.min))
        end_dt = timezone.make_aware(datetime.datetime.combine(yesterday, datetime.time.max))
    elif date_filter == 'custom' and from_date and to_date:
        try:
            f = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
            t = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()
            start_dt = timezone.make_aware(datetime.datetime.combine(f, datetime.time.min))
            end_dt = timezone.make_aware(datetime.datetime.combine(t, datetime.time.max))
        except Exception:
            pass

    if start_dt and end_dt:
        qs = qs.filter(created_at__gte=start_dt, created_at__lte=end_dt)

    # Get Order References
    filtered_refs = qs.exclude(order_reference__isnull=True).exclude(order_reference='').values_list('order_reference', flat=True).distinct()

    # Fetch Orders
    orders_qs = Order.objects.filter(order_number__in=filtered_refs).order_by('-created_at').prefetch_related('items')

    # Create Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Define Header Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")

    # Write Headers
    headers = [
        "Order ID",
        "Order Items",
        "Country",
        "City",
        "Order Status",
        "Total Amount",
        "Order Date"
    ]
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center

    # Write Data Rows
    row_num = 2
    for ord_obj in orders_qs:
        # Items
        item_names = [item.product_name or (item.product.name if item.product else "Unknown") for item in ord_obj.items.all()]
        items_str = ", ".join(item_names)
        
        # Country/City lookup
        cd = CheckoutDetails.objects.filter(order_reference=ord_obj.order_number).first()
        country_val = cd.country if cd else 'Unspecified'
        city_val = ord_obj.shipping_city or (cd.city if cd else 'Unspecified')

        # Status Display
        status_display = ord_obj.get_status_display()
        
        # Date
        date_str = ord_obj.created_at.strftime("%Y-%m-%d %H:%M") if ord_obj.created_at else ""

        row_data = [
            ord_obj.order_number,
            items_str,
            country_val,
            city_val,
            status_display,
            float(ord_obj.total_amount),
            date_str
        ]

        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = alignment_left
            # Format currency column as number
            if col_num == 6: # Total Amount
                cell.number_format = '0.00'

        row_num += 1

    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        # Cap width at 50 to prevent huge columns
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)


    # Determine filename
    filename = "orders_all_countries.xlsx"
    if selected_country == "US":
        filename = "orders_us.xlsx"
    elif selected_country == "CA":
        filename = "orders_ca.xlsx"

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
